import os
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

def format_sheet(ws):
    # Set column widths to fit the content
    for col in ws.columns:
        max_length = 0
        column = None
        for cell in col:
            if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                column = cell.column_letter  # Get the column name
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
        if column:
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

def write_titles(criteria_dict, titles, ws):

    # Create a list of titles with extra spaces for formatting
    title_row = [title + ' '*(len(title)-1) for title in titles]

    # Write the titles to the first row
    len_acc = 1
    for t in titles:
        criteria = criteria_dict[t]

        ws.merge_cells(start_row=1, start_column=len_acc, end_row=1, end_column=len_acc + len(criteria) - 1)
        ws.cell(row=1, column=len_acc, value=t).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=1, column=len_acc, value=t).font = Font(bold=True)

        # Write the criteria to the second row
        for i, criterion in enumerate(criteria):
            ws.cell(row=2, column=len_acc + i, value=criterion).alignment = Alignment(horizontal='center', vertical='center')
        
        # Set a thin border for borders of the criteria sets
        ws.cell(row=2, column=len_acc + i, value=criterion).border = Border(right=Side(style='thin'))

        len_acc += len(criteria)

# given a df and a worksheet, write the data to the worksheet
def write_data(df, ws):
    # Check if df is a DataFrame
    if not isinstance(df, pd.DataFrame):
        raise TypeError("The provided data is not a pandas DataFrame")

    # Determine the next empty row
    next_row = 3

    # Write DataFrame to the worksheet starting from the next empty row
    for row in dataframe_to_rows(df, index=False, header=False):
        for c_idx, value in enumerate(row):
            ws.cell(row=next_row, column=c_idx + 1, value=value)
        next_row += 1


def save(path, df):

    # Create a workbook and select the active worksheet
    wb = openpyxl.Workbook()
    ws = wb.active

    # Define the critera column labels
    causal_criteria = ['paper_no', 'year', 'coder', 'statement', 'no_occurrences', 
                    'section', 'causal', 'causal_type', 'valence', 'full_sentence']

    pieters_criteria = ['power', 'c_power', 'reliability', 'c_reliability', 'confounds', 
                        'd_methods', 'a_confounds', 'sens_analysis', 't_basis', 
                        'l_analysis', 'sem', 'sem_reliability', 'limitations', 
                        'limitations_reliability', 'limitations_confounds', 
                        'limitations_methods', 'limitations_analysis', 
                        'limitations_sem', 'limitations_sem_reliability']

    criteria_dict = {'Causal Language': causal_criteria, 'Pieters\' dimensions': pieters_criteria}
    titles = ['Causal Language', 'Pieters\' dimensions'] # Titles for the columns

    # Write the titles and criteria to the worksheet
    write_titles(criteria_dict, titles, ws)

    # Write the data to the worksheet
    write_data(df, ws)

    # Format the worksheet
    format_sheet(ws)

    # Save the workbook to a file
    wb.save(path)
    print(f"Data has been written to {path}")

#make it run for main
if __name__ == "__main__":
    # construct a path
    desktop = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
    path = os.path.join(desktop, 'formatted_data.xlsx')
    save(path, df)